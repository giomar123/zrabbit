#!/usr/bin/env python3.11
import openpyxl
import mysql.connector
import os
from dotenv import load_dotenv
from datetime import datetime

# Load environment variables
load_dotenv('/home/ubuntu/zrabbit_business/.env')

# Database connection
db = mysql.connector.connect(
    host=os.getenv('DATABASE_URL').split('@')[1].split('/')[0].split(':')[0],
    user=os.getenv('DATABASE_URL').split('://')[1].split(':')[0],
    password=os.getenv('DATABASE_URL').split(':')[2].split('@')[0],
    database=os.getenv('DATABASE_URL').split('/')[-1].split('?')[0]
)

cursor = db.cursor()

# Load Excel file with data_only=True to get calculated values instead of formulas
wb = openpyxl.load_workbook('/home/ubuntu/Downloads/12hmJhmOZbrOCRTGp7CX1jtmeysGW1j6qX42HSsI4Ip0.xlsx', data_only=True)

# Category mapping
category_map = {
    'Pokémon': 1,
    'Pokemon': 1,
    'Dragon Ball Z': 2,
    'Merch': 3,
    'One Piece': 4,
    'Yu Gi Oh': 5
}

def clean_price(value):
    """Clean price string and convert to float"""
    if value is None or value == '':
        return 0.0
    if isinstance(value, (int, float)):
        return float(value)
    # Remove S/., spaces, and commas
    cleaned = str(value).replace('S/.', '').replace('S/', '').replace(',', '').strip()
    try:
        return float(cleaned)
    except:
        return 0.0

def format_date(date_value):
    """Format date to YYYY-MM-DD"""
    if isinstance(date_value, datetime):
        return date_value.strftime('%Y-%m-%d')
    if isinstance(date_value, str):
        # Try parsing different date formats
        for fmt in ['%d/%m/%Y', '%Y-%m-%d', '%d-%m-%Y']:
            try:
                dt = datetime.strptime(date_value, fmt)
                return dt.strftime('%Y-%m-%d')
            except:
                continue
    return str(date_value)

print("=== Importing Products from 'Listas' sheet ===")
ws_listas = wb['Listas']
products_imported = 0
product_map = {}  # Map code -> id

for row in ws_listas.iter_rows(min_row=1, values_only=True):
    if not row[0]:  # Skip empty rows
        continue
    
    code = str(row[0]).strip()
    name = str(row[1]).strip() if row[1] else ''
    
    if not code or not name:
        continue
    
    # Extract category from code prefix
    prefix = code[:3]
    category_id = None
    if prefix == 'POK':
        category_id = 1
    elif prefix == 'DBZ':
        category_id = 2
    elif prefix == 'MRC':
        category_id = 3
    elif prefix == 'OPI':
        category_id = 4
    elif prefix == 'YGO':
        category_id = 5
    
    if category_id:
        cursor.execute(
            "INSERT INTO products (code, name, categoryId) VALUES (%s, %s, %s)",
            (code, name, category_id)
        )
        product_id = cursor.lastrowid
        product_map[code] = product_id
        products_imported += 1

db.commit()
print(f"✅ Imported {products_imported} products")

print("\n=== Importing Purchases from 'Compras' sheet ===")
ws_compras = wb['Compras']
purchases_imported = 0

for row in ws_compras.iter_rows(min_row=2, values_only=True):
    if not row[0]:  # Skip empty rows
        continue
    
    purchase_date = format_date(row[0])
    product_code = str(row[1]).strip() if row[1] else ''
    quantity = int(row[4]) if row[4] else 0
    unit_price = clean_price(row[5])
    total = clean_price(row[6])
    status = str(row[8]).strip() if row[8] else 'Recibido'
    detail = str(row[9]).strip() if row[9] else ''
    
    if not product_code or product_code not in product_map:
        continue
    
    product_id = product_map[product_code]
    
    # Calculate suggested price with 30% margin
    suggested_price = unit_price * 1.30
    
    cursor.execute(
        "INSERT INTO purchases (purchaseDate, productId, quantity, unitPrice, total, suggestedPrice, status, detail) VALUES (%s, %s, %s, %s, %s, %s, %s, %s)",
        (purchase_date, product_id, quantity, unit_price, total, suggested_price, status, detail)
    )
    purchases_imported += 1

db.commit()
print(f"✅ Imported {purchases_imported} purchases")

print("\n=== Importing Sales from 'Ventas' sheet ===")
ws_ventas = wb['Ventas']
sales_imported = 0

for row in ws_ventas.iter_rows(min_row=2, values_only=True):
    if not row[0]:  # Skip empty rows
        continue
    
    sale_date = format_date(row[0])
    product_code = str(row[1]).strip() if row[1] else ''
    quantity = int(row[3]) if row[3] else 0
    unit_price = clean_price(row[4])
    total = clean_price(row[5])
    buyer_name = str(row[6]).strip() if row[6] else ''
    buyer_email = str(row[7]).strip() if row[7] else ''
    buyer_phone = str(row[8]).strip() if row[8] else ''
    
    if not product_code or product_code not in product_map:
        continue
    
    product_id = product_map[product_code]
    
    cursor.execute(
        "INSERT INTO sales (saleDate, productId, quantity, unitPrice, total, buyerName, buyerEmail, buyerPhone) VALUES (%s, %s, %s, %s, %s, %s, %s, %s)",
        (sale_date, product_id, quantity, unit_price, total, buyer_name, buyer_email, buyer_phone)
    )
    sales_imported += 1

db.commit()
print(f"✅ Imported {sales_imported} sales")

print("\n=== Importing Investments from 'Inversiones' sheet ===")
ws_inversiones = wb['Inversiones']
investments_imported = 0

for row in ws_inversiones.iter_rows(min_row=2, values_only=True):
    if not row[0]:  # Skip empty rows
        continue
    
    investment_date = format_date(row[0])
    description = str(row[1]).strip() if row[1] else ''
    investor = str(row[2]).strip() if row[2] else ''
    amount = clean_price(row[3])
    
    if not investor or amount == 0:
        continue
    
    cursor.execute(
        "INSERT INTO investments (investmentDate, description, investor, amount) VALUES (%s, %s, %s, %s)",
        (investment_date, description, investor, amount)
    )
    investments_imported += 1

db.commit()
print(f"✅ Imported {investments_imported} investments")

print("\n=== Verifying Totals ===")
cursor.execute("SELECT SUM(CAST(total AS DECIMAL(10,2))) FROM sales WHERE saleDate LIKE '2025-12%'")
dec_sales = cursor.fetchone()[0] or 0
print(f"December 2025 Sales: S/. {dec_sales:.2f}")

cursor.execute("SELECT SUM(CAST(total AS DECIMAL(10,2))) FROM purchases WHERE purchaseDate LIKE '2025-12%'")
dec_purchases = cursor.fetchone()[0] or 0
print(f"December 2025 Purchases: S/. {dec_purchases:.2f}")

cursor.execute("SELECT SUM(CAST(amount AS DECIMAL(10,2))) FROM investments")
total_investments = cursor.fetchone()[0] or 0
print(f"Total Investments: S/. {total_investments:.2f}")

cursor.close()
db.close()

print("\n✅ Import completed successfully!")
