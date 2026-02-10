#!/usr/bin/env python3
import openpyxl
import mysql.connector
from datetime import datetime
import os
from dotenv import load_dotenv

load_dotenv()

# Parse DATABASE_URL
db_url = os.getenv("DATABASE_URL")
# Format: mysql://user:pass@host:port/dbname
parts = db_url.replace("mysql://", "").split("@")
user_pass = parts[0].split(":")
host_port_db = parts[1].split("/")
host_port = host_port_db[0].split(":")

conn = mysql.connector.connect(
    host=host_port[0],
    port=int(host_port[1]) if len(host_port) > 1 else 3306,
    user=user_pass[0],
    password=user_pass[1],
    database=host_port_db[1].split("?")[0],
    ssl_disabled=False
)

cursor = conn.cursor(dictionary=True)

print("Starting Excel import...\n")

# Load workbook
wb = openpyxl.load_workbook('/home/ubuntu/upload/VentadeFigurasv2(1).xlsx')

# ============= IMPORT PRODUCTS FROM "Listas" SHEET =============
print("📦 Importing products...")
ws_listas = wb["Listas"]

# Get category mapping
cursor.execute("SELECT id, code FROM categories")
category_map = {row['code']: row['id'] for row in cursor.fetchall()}

products_imported = 0
# Start from row 2, skip header
for row_idx in range(2, ws_listas.max_row + 1):
    codigo = ws_listas.cell(row=row_idx, column=1).value
    nombre = ws_listas.cell(row=row_idx, column=2).value
    
    if not codigo or not nombre:
        continue
    
    # Extract category code (first 3 letters)
    category_code = str(codigo)[:3]
    
    if category_code not in category_map:
        print(f"  ⚠️  Category not found for code: {category_code}")
        continue
    
    category_id = category_map[category_code]
    
    try:
        cursor.execute(
            "INSERT INTO products (code, name, categoryId) VALUES (%s, %s, %s)",
            (codigo, nombre, category_id)
        )
        products_imported += 1
    except mysql.connector.IntegrityError:
        print(f"  ⚠️  Product already exists: {codigo}")
    except Exception as e:
        print(f"  ❌ Error importing product {codigo}: {e}")

conn.commit()
print(f"✓ Products imported: {products_imported}\n")

# Get product mapping
cursor.execute("SELECT id, code FROM products")
product_map = {row['code']: row['id'] for row in cursor.fetchall()}

# ============= IMPORT PURCHASES FROM "Compras" SHEET =============
print("🛒 Importing purchases...")
ws_compras = wb["Compras"]

purchases_imported = 0
# Headers in row 1, data starts from row 2
for row_idx in range(2, min(ws_compras.max_row + 1, 1000)):  # Limit to first 1000 rows
    fecha = ws_compras.cell(row=row_idx, column=1).value
    codigo = ws_compras.cell(row=row_idx, column=2).value
    cantidad = ws_compras.cell(row=row_idx, column=5).value
    precio_unitario = ws_compras.cell(row=row_idx, column=6).value
    total = ws_compras.cell(row=row_idx, column=7).value
    precio_sugerido = ws_compras.cell(row=row_idx, column=8).value
    estado = ws_compras.cell(row=row_idx, column=9).value
    detalle = ws_compras.cell(row=row_idx, column=10).value
    
    if not fecha or not codigo:
        continue
    
    if codigo not in product_map:
        print(f"  ⚠️  Product not found: {codigo}")
        continue
    
    product_id = product_map[codigo]
    
    # Convert date
    if isinstance(fecha, datetime):
        date_str = fecha.strftime("%Y-%m-%d")
    else:
        date_str = str(fecha)
    
    try:
        cursor.execute(
            """INSERT INTO purchases 
               (purchaseDate, productId, quantity, unitPrice, total, suggestedPrice, status, detail) 
               VALUES (%s, %s, %s, %s, %s, %s, %s, %s)""",
            (date_str, product_id, cantidad or 0, precio_unitario or 0, total or 0, 
             precio_sugerido or 0, estado or "Recibido", detalle or "")
        )
        purchases_imported += 1
    except Exception as e:
        print(f"  ❌ Error importing purchase: {e}")

conn.commit()
print(f"✓ Purchases imported: {purchases_imported}\n")

# ============= IMPORT SALES FROM "Ventas" SHEET =============
print("💰 Importing sales...")
ws_ventas = wb["Ventas"]

sales_imported = 0
# Headers in row 1, actual headers in row 2, data starts from row 3
for row_idx in range(3, min(ws_ventas.max_row + 1, 1000)):  # Limit to first 1000 rows
    fecha = ws_ventas.cell(row=row_idx, column=1).value
    codigo = ws_ventas.cell(row=row_idx, column=2).value
    cantidad = ws_ventas.cell(row=row_idx, column=4).value
    precio_unitario = ws_ventas.cell(row=row_idx, column=5).value
    total = ws_ventas.cell(row=row_idx, column=6).value
    comprador = ws_ventas.cell(row=row_idx, column=7).value
    correo = ws_ventas.cell(row=row_idx, column=8).value
    telefono = ws_ventas.cell(row=row_idx, column=9).value
    
    if not fecha or not codigo:
        continue
    
    if codigo not in product_map:
        print(f"  ⚠️  Product not found: {codigo}")
        continue
    
    product_id = product_map[codigo]
    
    # Convert date
    if isinstance(fecha, datetime):
        date_str = fecha.strftime("%Y-%m-%d")
    else:
        date_str = str(fecha)
    
    try:
        cursor.execute(
            """INSERT INTO sales 
               (saleDate, productId, quantity, unitPrice, total, buyerName, buyerEmail, buyerPhone) 
               VALUES (%s, %s, %s, %s, %s, %s, %s, %s)""",
            (date_str, product_id, cantidad or 0, precio_unitario or 0, total or 0,
             comprador or "", correo or "", telefono or "")
        )
        sales_imported += 1
    except Exception as e:
        print(f"  ❌ Error importing sale: {e}")

conn.commit()
print(f"✓ Sales imported: {sales_imported}\n")

# ============= IMPORT INVESTMENTS FROM "Inversiones" SHEET =============
print("💵 Importing investments...")
ws_inversiones = wb["Inversiones"]

investments_imported = 0
# Headers in row 1, data starts from row 2
for row_idx in range(2, min(ws_inversiones.max_row + 1, 1000)):  # Limit to first 1000 rows
    fecha = ws_inversiones.cell(row=row_idx, column=1).value
    descripcion = ws_inversiones.cell(row=row_idx, column=2).value
    inversor = ws_inversiones.cell(row=row_idx, column=3).value
    monto = ws_inversiones.cell(row=row_idx, column=4).value
    
    if not fecha or not inversor or not monto:
        continue
    
    # Convert date
    if isinstance(fecha, datetime):
        date_str = fecha.strftime("%Y-%m-%d")
    else:
        date_str = str(fecha)
    
    try:
        cursor.execute(
            """INSERT INTO investments 
               (investmentDate, description, investor, amount) 
               VALUES (%s, %s, %s, %s)""",
            (date_str, descripcion or "", inversor, monto or 0)
        )
        investments_imported += 1
    except Exception as e:
        print(f"  ❌ Error importing investment: {e}")

conn.commit()
print(f"✓ Investments imported: {investments_imported}\n")

cursor.close()
conn.close()

print("✅ Excel import completed successfully!")
print(f"\nSummary:")
print(f"  - Products: {products_imported}")
print(f"  - Purchases: {purchases_imported}")
print(f"  - Sales: {sales_imported}")
print(f"  - Investments: {investments_imported}")
